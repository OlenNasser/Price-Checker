import urllib.request
from bs4 import BeautifulSoup
from openpyxl import Workbook
import openpyxl as opxl
import os
from datetime import date
import datetime
current_date = datetime.datetime.now()


class MAIN:
    def __init__(self):
        self.url = URL()
        self.html = HTML()
        self.excel = EXCEL()
        self.text = TEXT()
        self.convert = CONVERSIONS()
        pass

    def main(self):
        self.excel.create_excel()
        self.excel.write_excel(self.convert.convert_date_to_int(), self.html.get_span(self.url.get_html()))
        pass


class URL:
    def __init__(self):
        self.url = "https://www.cardmarket.com/en/Pokemon/Products/Elite-Trainer-Boxes/Celebrations-Elite-Trainer-Box"
        pass

    def main(self):

        pass
    def get_html(self):
        self.content = urllib.request.urlopen(self.url).read()
        return self.content
        pass
    
class HTML:
    def __init__(self):
        pass

    def main(self):
        pass
    def get_span(self, content):
        soup = BeautifulSoup(content,'html.parser')
        s_class = soup.find_all(class_="col-6 col-xl-7")
        span_tings=[]
        for i in s_class:
            if i.find('span') != None:
                span_tings.append(i.find('span').text)

        days_price = str(span_tings[3])
        cdays_price = days_price.replace('€','')
        bdays_price = cdays_price.replace(',','.')
        return float(bdays_price)*1.67
        
        
class EXCEL:
    def __init__(self):
        pass

    def main(self):
        pass
    def create_excel(self):
        if main.text.check_exists('prices.xlsx') == False:
            wb = Workbook() 
            ws = wb.active
            ws.title = "prices"
            ws['A1'] = "date"
            ws['B1'] = "price"
            wb.save('prices.xlsx')
        else:
            pass
            
        pass
    def write_excel(self, date, price):
        wb = opxl.load_workbook('prices.xlsx')
        ws = wb['prices']
        ws.append([date, price])
        wb.save('prices.xlsx')

class TEXT:
    def __init__(self):
        pass
    def main(self):
        pass
    def open_read_file(self, name):
        pass
    def check_exists(self, name):
        if os.path.exists(name):
            return True
        else:
            return False
class CONVERSIONS:
    def __init__(self):
        pass
    def main(self):
        pass
    def convert_date_to_int(self):
        n = int(current_date.year*10000 + current_date.month * 100 + current_date.day)
        return n
        
        pass
    def convert_to_string(self, int):
        pass
if __name__ == '__main__':
    main = MAIN()
    main.main()
    